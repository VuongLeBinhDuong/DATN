"""Turn spreadsheet rows into plain text lines for the same lab parser as PDF."""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

_FLOATISH = re.compile(r"^[+-]?[0-9]+([.,][0-9]+)?([eE][+-]?[0-9]+)?$")

# Bảng phiếu XN kiểu: | Xét nghiệm | Kết quả | Giá trị BT | Đơn vị | Máy |
_LAB_COLS = 6

# Kết quả định tính (vd. NC tiểu: cột B trống, NEGATIVE nằm cột C)
_QUAL_TOKENS = re.compile(
    r"(?i)^(negative|positive|trace|reactive|non[-\s]?reactive|"
    r"normal|abnormal|present|absent|clear|cloudy|neg\.?|pos\.?)(\b|$)"
)
_QUAL_ANY = re.compile(r"(?i)\b(negative|positive|trace|reactive)\b")


def _looks_qualitative_lab_result(s: str) -> bool:
    t = (s or "").strip()
    if not t:
        return False
    if _QUAL_TOKENS.match(t.split()[0] if t else ""):
        return True
    if re.match(r"^[\+\-]{1,3}$", t.strip()):
        return True
    if _QUAL_ANY.search(t):
        return True
    return False


def _looks_like_reference_only_cell(s: str) -> bool:
    """Ô chỉ chứa khoảng tham chiếu (số / Nam Nữ), không phải KQ định tính."""
    t = (s or "").strip()
    if not t:
        return False
    if _looks_qualitative_lab_result(t):
        return False
    if re.search(r"(?i)\bnam:|nữ:|nam\s*<|nữ\s*<", t):
        return True
    # chỉ số, dấu so sánh, gạch — giống dòng tham chiếu máu
    if re.match(r"^[\d\s.,<≥≤%\-–/]+$", t) and re.search(r"\d", t):
        return True
    return False


def _effective_lab_result_fields(cells: list[str]) -> tuple[str, str, str, str]:
    """
    Trả về (result, ref, unit, máy).

    Chuẩn sinh hóa: B=Kết quả, C=tham chiếu, D=đơn vị, E=máy.
    Một số phiếu (vd. NC tiểu): B trống, NEGATIVE ở C, D=Labumat → coi C là Kết quả.
    """
    c1 = cells[1].strip() if len(cells) > 1 else ""
    c2 = cells[2].strip().replace("\n", " ") if len(cells) > 2 else ""
    c3 = cells[3].strip() if len(cells) > 3 else ""
    c4 = cells[4].strip() if len(cells) > 4 else ""
    c5 = cells[5].strip() if len(cells) > 5 else ""

    if c1:
        return c1, c2, c3, c4

    if not c2:
        return "", c2, c3, c4

    if _looks_qualitative_lab_result(c2):
        return c2, "", "", c3

    if not _looks_like_reference_only_cell(c2):
        # kết quả lệch sang cột C (hiếm), dồn phần sau
        return c2, c3, c4, c5

    return "", c2, c3, c4


def _row_cells_padded(row: tuple[Any, ...], *, min_cols: int = _LAB_COLS) -> list[str]:
    """Giữ cột cố định (kể cả ô trống) để không lệch Kết quả / Tham chiếu."""
    cells = [_norm_cell(v) for v in row]
    while len(cells) < min_cols:
        cells.append("")
    return cells[: max(min_cols, len(cells))]


def _row_looks_like_lab_header(cells: list[str]) -> bool:
    """Tránh nhầm với tiêu đề 'PHIẾU KẾT QUẢ XÉT NGHIỆM' (một ô)."""
    c0 = cells[0].strip()
    c1 = (cells[1].strip() if len(cells) > 1 else "") or ""
    return c0 == "XÉT NGHIỆM" and "KẾT QUẢ" in c1


def _looks_like_lab_test_label(label: str) -> bool:
    """Loại dòng thông tin hành chính trên cùng sheet."""
    if len(label) > 55:
        return False
    low = label.lower()
    skip = (
        "họ tên",
        "bác sĩ",
        "bác sỹ",
        "chẩn đoán",
        "ngày xét",
        "mã code",
        "địa chỉ",
        "phiếu kết quả",
        "bệnh viện",
        "bv ",
    )
    return not any(s in low for s in skip)


def _format_lab_table_row(cells: list[str]) -> str | None:
    """Một dòng xét nghiệm: kết quả tách bạch khỏi tham chiếu (tránh LLM đọc nhầm)."""
    label = cells[0].strip()
    if not label or len(label) > 100:
        return None
    if _row_looks_like_lab_header(cells):
        return None
    if not _looks_like_lab_test_label(label):
        return None
    result, ref, unit, mach = _effective_lab_result_fields(cells[:_LAB_COLS])
    ref = ref.replace("\n", " ").replace("  ", " ").strip()
    if not result and not ref and not unit:
        return None
    # Không trích dòng hoàn toàn trống Kết quả (sau khi đã tính NEGATIVE ở cột kế)
    if not result:
        return None
    res_disp = result.replace(",", ".")
    parts = [f"{label}: {res_disp}"]
    if unit:
        parts.append(f"Đơn vị: {unit}")
    if ref:
        parts.append(f"Tham chiếu in trên phiếu (không phải kết quả bệnh nhân): {ref}")
    if mach and mach not in (unit,):
        parts.append(f"Máy: {mach}")
    return " | ".join(parts)


def _is_floatish(s: str) -> bool:
    t = s.strip().replace(",", ".")
    if not t:
        return False
    return bool(_FLOATISH.match(t))


def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _include_hidden_xlsx_rows() -> bool:
    return os.getenv("MEDICAL_RECORD_XLSX_INCLUDE_HIDDEN", "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )


def _is_row_hidden(ws: Worksheet, row_idx_1based: int) -> bool:
    """Excel có thể ẩn hàng (template / chỉ số không dùng); mặc định bỏ qua khi trích."""
    dim = ws.row_dimensions.get(row_idx_1based)
    if dim is None:
        return False
    return bool(getattr(dim, "hidden", False))


def _skip_lab_row_no_result(padded: list[str]) -> bool:
    """Trong bảng lab: bỏ dòng không có Kết quả (kể cả NEGATIVE ở cột C khi B trống)."""
    label = padded[0].strip() if padded else ""
    if not label or _row_looks_like_lab_header(padded):
        return False
    if not _looks_like_lab_test_label(label):
        return False
    result, _, _, _ = _effective_lab_result_fields(padded[:_LAB_COLS])
    return not result


def _raw_row_tsv(row: tuple[Any, ...]) -> str | None:
    """Một dòng TSV (tab): giá trị ô theo thứ tự cột; bỏ ô trống cuối hàng."""
    cells = [_norm_cell(v) for v in row]
    while cells and cells[-1] == "":
        cells.pop()
    if not any(c.strip() for c in cells):
        return None
    return "\t".join(cells)


def extract_raw_text_from_xlsx(
    path: Path | str,
    *,
    sheet_name: str | None = None,
) -> tuple[str, dict[str, Any]]:
    """
    Dump “thô” theo lưới Excel: mỗi hàng không trống = một dòng, ô cách nhau bằng tab.

    Không gộp nhãn LLM, không chuẩn hóa cột Kết quả/tham chiếu — dùng để đối chiếu với file gốc.
    Chính sách hàng ẩn giống :func:`extract_text_from_xlsx` (biến môi trường
    ``MEDICAL_RECORD_XLSX_INCLUDE_HIDDEN``).
    """
    p = Path(path)
    wb = load_workbook(p, read_only=False, data_only=True)
    try:
        names = [sheet_name] if sheet_name else list(wb.sheetnames)
        if sheet_name and sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name!r}; available: {wb.sheetnames}")
        parts: list[str] = []
        row_counts: list[int] = []
        include_hidden = _include_hidden_xlsx_rows()
        for name in names:
            ws = wb[name]
            lines: list[str] = []
            n = 0
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, values_only=True),
                start=1,
            ):
                if not include_hidden and _is_row_hidden(ws, row_idx):
                    continue
                line = _raw_row_tsv(row)
                if line is None:
                    continue
                n += 1
                lines.append(line)
            row_counts.append(n)
            if lines:
                parts.append(f"--- Sheet: {name} ---\n" + "\n".join(lines))
        text = "\n\n".join(parts)
        return text, {
            "format": "xlsx_raw",
            "sheets_read": names,
            "row_counts_per_sheet": row_counts,
            "skip_hidden_rows": not include_hidden,
            "delimiter": "tab",
        }
    finally:
        wb.close()


def extract_text_from_xlsx(
    path: Path | str,
    *,
    sheet_name: str | None = None,
) -> tuple[str, dict[str, Any]]:
    """
    Flatten workbook to lines: ``Label: value unit`` when column 2 looks numeric.
    If sheet_name is absent, all sheets are concatenated.
    """
    p = Path(path)
    # read_only=False: cần row_dimensions.hidden để khớp những gì user nhìn thấy trên Excel.
    wb = load_workbook(p, read_only=False, data_only=True)
    try:
        names = [sheet_name] if sheet_name else list(wb.sheetnames)
        if sheet_name and sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name!r}; available: {wb.sheetnames}")
        parts: list[str] = []
        row_counts: list[int] = []
        include_hidden = _include_hidden_xlsx_rows()
        for name in names:
            ws = wb[name]
            lines: list[str] = []
            n = 0
            in_lab_table = False
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, values_only=True),
                start=1,
            ):
                if not include_hidden and _is_row_hidden(ws, row_idx):
                    continue
                padded = _row_cells_padded(row)
                if not any(x.strip() for x in padded):
                    continue
                n += 1
                if _row_looks_like_lab_header(padded):
                    in_lab_table = True
                    compact = [x for x in padded[:5] if x.strip()]
                    lines.append(" | ".join(compact))
                    continue
                if in_lab_table and padded[0].strip():
                    formatted = _format_lab_table_row(padded[:_LAB_COLS])
                    if formatted:
                        lines.append(formatted)
                        continue
                    if _skip_lab_row_no_result(padded):
                        continue
                cells = [s for v in row if (s := _norm_cell(v))]
                if not cells:
                    continue
                if len(cells) >= 2 and _is_floatish(cells[1]):
                    val = cells[1].strip().replace(",", ".")
                    extra = f" {cells[2]}" if len(cells) > 2 else ""
                    lines.append(f"{cells[0]}: {val}{extra}")
                else:
                    lines.append(" | ".join(cells))
            row_counts.append(n)
            if lines:
                parts.append(f"--- Sheet: {name} ---\n" + "\n".join(lines))
        text = "\n\n".join(parts)
        return text, {
            "format": "xlsx",
            "sheets_read": names,
            "row_counts_per_sheet": row_counts,
            "skip_hidden_rows": not include_hidden,
            "skip_empty_lab_result": True,
        }
    finally:
        wb.close()


def extract_structured_rows_from_xlsx(
    path: Path | str,
    *,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """
    Preserve each row as a list of cell strings (ragged rows). Trailing empty cells dropped per row.
    Empty rows skipped. Use this when you need table structure, not the flattened text used for lab parsing.
    """
    p = Path(path)
    wb = load_workbook(p, read_only=False, data_only=True)
    try:
        names = [sheet_name] if sheet_name else list(wb.sheetnames)
        if sheet_name and sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name!r}; available: {wb.sheetnames}")
        include_hidden = _include_hidden_xlsx_rows()
        sheets: list[dict[str, Any]] = []
        for name in names:
            ws = wb[name]
            rows: list[list[str]] = []
            in_lab_table = False
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, values_only=True),
                start=1,
            ):
                if not include_hidden and _is_row_hidden(ws, row_idx):
                    continue
                padded = _row_cells_padded(row)
                if not any(x.strip() for x in padded):
                    continue
                if _row_looks_like_lab_header(padded):
                    in_lab_table = True
                elif in_lab_table and _skip_lab_row_no_result(padded):
                    continue
                cells = [_norm_cell(v) if v is not None else "" for v in row]
                while cells and cells[-1] == "":
                    cells.pop()
                if not any(cells):
                    continue
                rows.append(cells)
            sheets.append(
                {
                    "name": name,
                    "row_count": len(rows),
                    "rows": rows,
                    "skip_hidden_rows": not include_hidden,
                    "skip_empty_lab_result": True,
                }
            )
        return {
            "format": "xlsx_structured",
            "source_file": str(p.resolve()),
            "sheets": sheets,
        }
    finally:
        wb.close()
